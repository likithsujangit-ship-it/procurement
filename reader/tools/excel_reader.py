"""
Excel Spreadsheets (XLSX/XLS) Content Extraction Module.
Reads sheets and cells from Excel workbooks using openpyxl.
"""

from pathlib import Path
import openpyxl
from tools.utils import setup_logger

logger = setup_logger("excel_reader")


def extract_excel_text(filepath: Path) -> str:
    """
    Extracts structured sheet names and cell tables from a .xlsx file.
    Provides troubleshooting fallback for older .xls files.
    
    Args:
        filepath: Path to the Excel spreadsheet on disk.
        
    Returns:
        Extracted text content as a string.
    """
    logger.info(f"Extracting sheets from Excel workbook: {filepath.name}")
    
    try:
        text_parts = []
        
        # Handle older .xls files using xlrd
        if filepath.suffix.lower() == ".xls":
            import xlrd
            wb = xlrd.open_workbook(filepath)
            for sheet in wb.sheets():
                text_parts.append(f"--- Sheet: {sheet.name} ---")
                rows_content = []
                for row_idx in range(sheet.nrows):
                    row = sheet.row_values(row_idx)
                    # Filter out completely empty rows
                    if not any(str(val).strip() for val in row):
                        continue
                    row_str = [str(val).strip() for val in row]
                    rows_content.append(" | ".join(row_str))
                
                if rows_content:
                    text_parts.append("\n".join(rows_content))
                else:
                    text_parts.append("[Empty Sheet]")
        else:
            # Handle modern .xlsx files using openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            for sheet_name in wb.sheetnames:
                text_parts.append(f"--- Sheet: {sheet_name} ---")
                sheet = wb[sheet_name]
                
                rows_content = []
                for row in sheet.iter_rows(values_only=True):
                    # Filter out completely empty rows
                    if not any(val is not None and str(val).strip() for val in row):
                        continue
                    row_str = [str(val).strip() if val is not None else "" for val in row]
                    rows_content.append(" | ".join(row_str))
                    
                if rows_content:
                    text_parts.append("\n".join(rows_content))
                else:
                    text_parts.append("[Empty Sheet]")
                    
        if not text_parts:
            return "[Excel file has no sheets]"
            
        return "\n\n".join(text_parts)

    except Exception as e:
        logger.error(f"Error reading Excel workbook {filepath.name}: {e}")
        return f"[Error reading Excel file: {e}]"
